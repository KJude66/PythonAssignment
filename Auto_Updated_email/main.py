'''
Name: Kum Jude Wung

Email: kum.wung@ictuniversity.edu.cm

iMatricule: ICTU20201127

Contact: 237654250309

Course: Programing in python 
    Python Programme to acces and update elements in an excel file
    *
    This project is a python based project that permits the access and updates of element in a database.
        The project was executed using the "openpyxl" pyhton libary which permits us to read and write excel files.add()
        it could be installed directly on command prompt using the command  " pip install openpyxl " for more visit "pip.org" website
        
    The programe recieves a created file with the extention .xlsx access its data and overide it as explained on each step down

'''

import openpyxl as xl

wb = xl.load_workbook('employees_data.xlsx')


sheet = wb['Sheet1']# sheet 1 is the active cell thats y it is called in the workbook(wb)

domain_name1 = '@helpinghands.cm'
domain_name2 = '@handsinhands.org'

# ------- CREATING THE EMAILS USING THE USER'S NAME ----------#

for row in range (2, sheet.max_row + 1):# begins from two because we're working with the second row elements

    cell = sheet.cell(row, 1 )# reads the element of column 1 in all the rows

    email = ((cell.value).lower() + domain_name1) # reads the element of the cell e.g jude and add @helpinghands.cm to it
      
    '''#-------- creating a new cell to store the email created --------#'''

    email_cell = sheet.cell(row, 3) # stores the  email in the 4th column(column C)

    ''''#--------- attributing the data to the new cells (column) created --------#'''
   
    email_cell.value = email
wb.save('employees_data.xlsx')

#------------ UPDATING THE EMAILS FOUND IN THE FILE ------------#

for row in range (2, sheet.max_row + 1):# begins from two because we're working with the second row elements
    
    cell = sheet.cell(row, 3)#telling the compiler to acces the element in R1 C3
    if domain_name1 in cell.value:
        new_email = (cell.value).replace(domain_name1, domain_name2)# interchanging the domain names
        
        sheet.cell(row, 3).value = new_email # tells the compiler that the new value of the cell in sheet1 is new email#
    
wb.save('employees_data.csv')

